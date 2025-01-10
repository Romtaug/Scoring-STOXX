receiver_email = "taugourdea@cy-tech.fr"

import re

def is_valid_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

if not is_valid_email(receiver_email):
    print(f"❌ Adresse email invalide : {receiver_email}")
    sys.exit(1)
    
#############################################################################

import subprocess
import sys

# Fonction pour installer un package si ce n'est pas déjà installé
def install_if_missing(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Liste des bibliothèques à vérifier et à installer si nécessaire
packages = [
    "requests",
    "beautifulsoup4",
    "yfinance",
    "pandas",
    "numpy",
    "pulp",
    "scikit-learn",
    "tqdm",
    "xlsxwriter",
    "pyperclip",
    "datetime"
]

# Installation des bibliothèques manquantes
for package in packages:
    install_if_missing(package)

# Import des bibliothèques nécessaires
import requests
import xlsxwriter
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
import numpy as np
from pulp import LpMaximize, LpProblem, LpVariable, lpSum
from sklearn.preprocessing import MinMaxScaler
from sklearn.linear_model import LinearRegression
from tqdm import tqdm
from datetime import datetime
import pyperclip
import re
import time
from IPython.display import display, HTML
import urllib.parse
from openpyxl import load_workbook

def adjust_column_width(file_path):
    """
    Ajuste la largeur des colonnes dans un fichier Excel en fonction du contenu affiché
    pour éviter une largeur excessive due aux URL complètes.
    """
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        # Ajuster uniquement en fonction du texte visible pour les colonnes de liens
                        if "Boursorama" in str(cell.value) or "Yahoo Finance" in str(cell.value):
                            visible_text_length = len("Boursorama")  # Ajusté au texte visible
                        else:
                            visible_text_length = len(str(cell.value))
                        max_length = max(max_length, visible_text_length)
                except Exception as e:
                    print(f"Erreur lors de l'ajustement de la colonne {col_letter}: {e}")
            adjusted_width = max_length + 1  # Ajouter une marge pour lisibilité
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)
    print(f"✅ Largeur des colonnes ajustée pour le fichier : {file_path}")

# Désactiver la barre de progression TQDM si besoin
tqdm.disable = False

print("Toutes les bibliothèques sont installées et importées avec succès.")

"""Scrapping des tickers des composants du SP500 sur Wikipédia et exportation des données de Yahoo Finance"""

# URL Wikipédia pour le S&P 500
url_sp500 = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"

# Fonction pour nettoyer les tickers
def clean_ticker(ticker):
    ticker = ticker.strip().upper()
    ticker = re.sub(r'\$', '', ticker)
    ticker = ticker.replace('.', '-')
    return ticker

# Scraper les tickers du S&P 500
def scrape_sp500_tickers(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find('table', {'id': 'constituents'})
    if not table:
        print("Tableau non trouvé sur Wikipédia.")
        return []

    tickers = [clean_ticker(row.find_all('td')[0].text) for row in table.find_all('tr')[1:]]
    return tickers

# Analyse des tickers avec yfinance
def analyze_sp500_tickers(tickers):
    analysis = {}
    failed_tickers = []

    for ticker in tqdm(tickers, desc="Récupération des tickers S&P 500"):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            history = stock.history(period="1d")
            close_price = history["Close"].iloc[0] if not history.empty else "N/A"

            analysis[ticker] = {
                "Ticker": ticker,
                "Name": info.get("shortName", "N/A"),
                "Sector": info.get("sector", "N/A"),
                "Industry": info.get("industry", "N/A"),
                "Country": info.get("country", "N/A"),
                "Market Cap": info.get("marketCap", "N/A"),
                "P/E Ratio": info.get("trailingPE", "N/A"),
                "Forward P/E": info.get("forwardPE", "N/A"),
                "Dividend Yield": info.get("dividendYield", "N/A") * 100 if info.get("dividendYield") else "N/A",
                "Revenue Growth": info.get("revenueGrowth", "N/A"),
                "EPS": info.get("trailingEps", "N/A"),
                "Beta": info.get("beta", "N/A"),
                "P/B Ratio": info.get("priceToBook", "N/A"),
                "Debt/Equity": info.get("debtToEquity", "N/A"),
                "Operating Cashflow": info.get("operatingCashflow", "N/A"),
                "Free Cashflow": info.get("freeCashflow", "N/A"),
                "Gross Margins": info.get("grossMargins", "N/A"),
                "Profit Margins": info.get("profitMargins", "N/A"),
                "Current Ratio": info.get("currentRatio", "N/A"),
                "Quick Ratio": info.get("quickRatio", "N/A"),
                "Recommendation": info.get("recommendationKey", "N/A"),
                "Target Mean Price": info.get("targetMeanPrice", "N/A"),
                "Previous Close": info.get("previousClose", "N/A"),
                "Open": info.get("open", "N/A"),
                "Day Low": info.get("dayLow", "N/A"),
                "Day High": info.get("dayHigh", "N/A"),
                "Volume": info.get("volume", "N/A"),
                "50 Day Average": info.get("fiftyDayAverage", "N/A"),
                "200 Day Average": info.get("twoHundredDayAverage", "N/A"),
                "Close Price": close_price,
            }
        except Exception as e:
            print(f"Erreur pour {ticker}: {e}")
            failed_tickers.append(ticker)
            analysis[ticker] = {"Error": str(e)}

        time.sleep(0.5)

    print(f"Tickers ayant échoué : {failed_tickers}")
    return analysis

#print("Récupération des données S&P 500 en cours...")
tickers_sp500 = scrape_sp500_tickers(url_sp500)
print(f"Nombre de données des tickers récupérés : {len(tickers_sp500)}")

tickers_sp500_sample = tickers_sp500

portfolio_analysis = analyze_sp500_tickers(tickers_sp500_sample)
df = pd.DataFrame.from_dict(portfolio_analysis, orient="index").fillna("N/A")

"""Télécharger le fichier Source"""

current_date = datetime.now().strftime("%Y-%m-%d")
file_path = f"Data-SP500_{current_date}.xlsx"
df.to_excel(file_path, index=False)
print(f"Fichier créé : {file_path}")
adjust_column_width(file_path)

"""Chargement des données historiques de cloture et mesure de potentiels créés"""

from tqdm import tqdm
import numpy as np
import pandas as pd
from sklearn.linear_model import LinearRegression
import yfinance as yf
from datetime import datetime

tqdm.disable = False

# Lecture du fichier
current_date = datetime.now().strftime("%Y-%m-%d")
file_path = f"Data-SP500_{current_date}.xlsx"
df = pd.read_excel(file_path)

# Vérification de la colonne 'Ticker'
if "Ticker" in df.columns:
    df = df.dropna(subset=["Ticker"])
    tickers = df["Ticker"].astype(str).unique()
else:
    raise ValueError("La colonne 'Ticker' n'existe pas dans le fichier.")

# Fonction pour analyser la tendance
def analyze_trend(ticker):
    try:
        stock = yf.Ticker(ticker)
        history = stock.history(period="max")

        if history.empty or len(history) < 2:
            return "Indéterminée", None, "Indéterminée"

        # Moyennes mobiles sur 6 mois et 12 mois
        history["SMA_6M"] = history["Close"].rolling(window=126, min_periods=1).mean()
        history["SMA_12M"] = history["Close"].rolling(window=252, min_periods=1).mean()
        trend = "Haussière" if history["SMA_6M"].iloc[-1] > history["SMA_12M"].iloc[-1] else "Baissière"

        # Régression linéaire pour estimer la pente
        history = history.reset_index()
        history["Days"] = np.arange(len(history))
        valid_data = history.dropna(subset=["Close", "Days"])

        model = LinearRegression()
        model.fit(valid_data[["Days"]], valid_data["Close"])
        slope = model.coef_[0]
        potential_growth = "Oui" if slope > 0 else "Non"

        return trend, slope, potential_growth
    except Exception:
        return "Indéterminée", None, "Indéterminée"

# Analyse des tickers avec progression
print("Analyse des tendances et prévisions...")
results = [analyze_trend(ticker) for ticker in tqdm(tickers, desc="Analyse des tickers", leave=False)]

# Création d'un DataFrame des résultats
df_results = pd.DataFrame(results, columns=["Tendance", "Pente", "Croissance Potentielle"], index=tickers)

# Fusion des résultats avec le DataFrame original
df = df.merge(df_results, left_on="Ticker", right_index=True, how="left")

# Fonctions pour convertir en scores
def trend_to_score(t):
    return 10 if t == "Haussière" else 1 if t == "Baissière" else np.nan

def growth_to_numeric(g):
    return 10 if g == "Oui" else 1 if g == "Non" else np.nan

def slope_to_score(s):
    return 10 if s and s > 0 else 1 if s is not None else np.nan

def target_potential(row):
    try:
        target = float(row["Target Mean Price"]) if row["Target Mean Price"] != "N/A" else np.nan
        closep = float(row["Close Price"]) if row["Close Price"] != "N/A" else np.nan
        if not pd.isna(target) and not pd.isna(closep) and closep != 0:
            return (target - closep) / closep * 100
    except:
        pass
    return np.nan

# Application des scores
df["Trend Score"] = df["Tendance"].apply(trend_to_score)
df["Potential Growth Numeric"] = df["Croissance Potentielle"].apply(growth_to_numeric)
df["Regression Score"] = df["Pente"].apply(slope_to_score)

df["Potential Calculated"] = (
    df["Regression Score"] *
    df["Potential Growth Numeric"] *
    df["Trend Score"]
)

df["Target Potential"] = df.apply(target_potential, axis=1)
# Calcul du total gain potentiel

df["Potential Calculated"] = df["Potential Calculated"].round(2)
df["Target Potential"] = df["Target Potential"].round(2)

#################################################################################

import pandas as pd
from datetime import datetime

# Export du fichier Excel initial
current_date = datetime.now().strftime('%Y-%m-%d')
output_file = f"Analysis-SP500_{current_date}.xlsx"
df.to_excel(output_file, index=False)
print(f"Calculs terminés avec succès. Fichier exporté : {output_file}")
adjust_column_width(output_file)

# Création d'une copie de df
df_cleaned = df.copy()

# Remplacement des valeurs non numériques et des valeurs manquantes par 0 uniquement pour les colonnes cibles
df_cleaned["Dividend Yield"] = df_cleaned["Dividend Yield"].apply(
    lambda x: 0 if pd.isnull(x) or (isinstance(x, str) and not x.replace(".", "").isdigit()) else float(x)
)
df_cleaned["Target Potential"] = df_cleaned["Target Potential"].apply(
    lambda x: 0 if pd.isnull(x) or (isinstance(x, str) and not x.replace(".", "").isdigit()) else float(x)
)

# Arrondi de la colonne "Close Price"
df_cleaned["Close Price"] = df_cleaned["Close Price"].round(2)

# Calcul du Total Gain Potential
df_cleaned["Total Gain Potential"] = df_cleaned["Dividend Yield"] + df_cleaned["Target Potential"]

# Arrondi des colonnes numériques pour un meilleur affichage
numeric_columns = ["Dividend Yield", "Target Potential", "Total Gain Potential"]
df_cleaned[numeric_columns] = df_cleaned[numeric_columns].round(2)

# Renommer les colonnes pour indiquer les pourcentages
df_cleaned.rename(
    columns={
        "Dividend Yield": "Dividend Yield (%)",
        "Target Potential": "Target Potential (%)",
        "Total Gain Potential": "Total Gain Potential (%)"
    },
    inplace=True
)

# Génération des liens avec des noms uniformes
df_cleaned["Boursorama Link"] = df_cleaned["Ticker"].apply(
    lambda ticker: f"https://www.boursorama.com/cours/{ticker.lower()}/"
)
df_cleaned["Yahoo Link"] = df_cleaned["Ticker"].apply(
    lambda ticker: f"https://finance.yahoo.com/quote/{ticker}?p={ticker}"
)

# Sélection des colonnes importantes
columns_to_include = [
    "Ticker",
    "Name",
    "Sector",
    "Close Price",
    "Dividend Yield (%)",
    "Target Potential (%)",
    "Total Gain Potential (%)",
    "Boursorama Link",
    "Yahoo Link"
]

# Tri des données pour chaque feuille
df_sorted_by_dividend = df_cleaned.sort_values(by="Dividend Yield (%)", ascending=False)
df_sorted_by_potential = df_cleaned.sort_values(by="Target Potential (%)", ascending=False)
df_sorted_by_total_gain = df_cleaned.sort_values(by="Total Gain Potential (%)", ascending=False)

# Chemin du fichier de sortie final
output_file = f"SP500_Ranking_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

# Création du fichier Excel avec 3 feuilles et liens cliquables
with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
    for sheet_name, df_sorted in {
        "By Total Gain": df_sorted_by_total_gain,
        "By Potential": df_sorted_by_potential,
        "By Dividend": df_sorted_by_dividend,
    }.items():
        # Export des colonnes importantes seulement
        df_export = df_sorted[columns_to_include]
        df_export.to_excel(writer, sheet_name=sheet_name, index=False)

        # Ajout des liens cliquables
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        link_format = workbook.add_format({"font_color": "blue", "underline": 1})

        for row_num, row in enumerate(df_export.itertuples(index=False), start=1):
            worksheet.write_url(
                row_num, df_export.columns.get_loc("Boursorama Link"),
                row[df_export.columns.get_loc("Boursorama Link")], link_format, "Boursorama"
            )
            worksheet.write_url(
                row_num, df_export.columns.get_loc("Yahoo Link"),
                row[df_export.columns.get_loc("Yahoo Link")], link_format, "Yahoo Finance"
            )

print(f"Fichier Excel créé avec succès : {output_file}")
adjust_column_width(output_file)

"""On va normaliser de 1 à 10 pour effectuer le scoring"""

import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler

current_date = datetime.now().strftime("%Y-%m-%d")
file_path = f"Analysis-SP500_{current_date}.xlsx"
df = pd.read_excel(file_path)
adjust_column_width(file_path)


# Colonnes à normaliser
columns_to_normalize = [
    "Market Cap", "P/E Ratio", "Dividend Yield", "Revenue Growth", "Beta",
    "P/B Ratio", "Target Potential", "Potential Calculated"
]

# Colonnes où une valeur faible est meilleure
low_is_better = ["P/E Ratio", "Beta", "P/B Ratio"]

# Colonnes où une valeur élevée est meilleure
high_is_better = ["Market Cap", "Dividend Yield", "Revenue Growth", "Target Potential", "Potential Calculated"]

# Assurez-vous que les colonnes sont numériques
for col in columns_to_normalize:
    df[col] = pd.to_numeric(df[col], errors='coerce')

# Suppression des lignes avec NaN dans les colonnes critiques
df = df.dropna(subset=columns_to_normalize)
df.replace([np.inf, -np.inf], np.nan, inplace=True)

# Normalisation MinMaxScaler (plage entre 1 et 10)
scaler = MinMaxScaler(feature_range=(1, 10))

# Appliquer la normalisation pour chaque colonne
for col in columns_to_normalize:
    if col in low_is_better:
        # Inverser les scores pour les colonnes où faible est meilleur
        df[f"{col}_normalized"] = 10 - scaler.fit_transform(df[[col]])
    else:
        # Normalisation classique pour les colonnes où élevé est meilleur
        df[f"{col}_normalized"] = scaler.fit_transform(df[[col]])

# Affichage des résultats normalisés pour vérification
#print("Aperçu des données normalisées :")
#print(df[[col for col in df.columns if "normalized" in col]].head())

"""Normalisation des données : suppression de ceux pas dans les critères de scoring, normalisation de 1 à 10"""

from pulp import LpMaximize, LpProblem, LpVariable, lpSum

def generate_portfolio(years, df, specialized=False, profit=False):
    if profit:
        # Nouveau portefeuille "Profit" avec moitié Dividend Yield et moitié Target Potential
        weights = {"Dividend Yield_normalized": 0.5, "Target Potential_normalized": 0.5}
    elif specialized:
        # Trading (spécialisé)
        weights = {"Target Potential_normalized": 1.0}
    else:
        # Portefeuilles classiques selon les années
        if years <= 3:
            weights = {"Target Potential_normalized": 0.3, "Potential Calculated_normalized": 0.3,
                       "Beta_normalized": 0.1, "P/E Ratio_normalized": 0.1, "Revenue Growth_normalized": 0.1, "Dividend Yield_normalized": 0.1}
        elif 3 < years <= 7:
            weights = {"Target Potential_normalized": 0.1, "Potential Calculated_normalized": 0.1,
                       "Beta_normalized": 0.2, "P/E Ratio_normalized": 0.2, "Revenue Growth_normalized": 0.2, "Dividend Yield_normalized": 0.2}
        else:
            weights = {"Target Potential_normalized": 0.1, "Potential Calculated_normalized": 0.1,
                       "Beta_normalized": 0.1, "P/E Ratio_normalized": 0.1, "Revenue Growth_normalized": 0.3, "Dividend Yield_normalized": 0.3}

    # Assurer que les colonnes utilisées sont numériques
    for col in weights:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')  # Remplace les valeurs non numériques par NaN

    # Calculer le score final
    df["Final Score"] = df[list(weights.keys())].mul(weights, axis=1).sum(axis=1)

    # Trier et filtrer les lignes ayant des valeurs valides
    df_sorted = df.dropna(subset=["Final Score"]).sort_values(by="Final Score", ascending=False)

    # Problème d'optimisation
    problem = LpProblem("Portfolio_Optimization", LpMaximize)
    variables = {i: LpVariable(name=str(i), lowBound=0, upBound=1, cat="Continuous") for i in df_sorted.index}

    problem += lpSum(variables[i] * df_sorted.at[i, "Final Score"] for i in df_sorted.index)
    problem += lpSum(variables[i] for i in df_sorted.index) == 1

    # Contraintes par secteur
    max_sector_proportion = 0.2  # Maximum 20% par secteur
    min_sector_proportion = 0.05  # Minimum 5% par secteur
    sectors = df_sorted["Sector"].unique()

    for sector in sectors:
        tickers_in_sector = df_sorted[df_sorted["Sector"] == sector].index
        problem += lpSum(variables[i] for i in tickers_in_sector) <= max_sector_proportion, f"Max_Sector_{sector}"
        problem += lpSum(variables[i] for i in tickers_in_sector) >= min_sector_proportion, f"Min_Sector_{sector}"

    # Résoudre le problème
    problem.solve()

    # Récupérer les résultats
    selected_tickers = [i for i in variables if variables[i].value() > 0]
    weights = {i: variables[i].value() for i in selected_tickers}
    df_selected = df_sorted.loc[selected_tickers].copy()
    df_selected["Weight"] = df_selected.index.map(weights)

    return df_selected

# Génération des portefeuilles
portfolios = {
    "Trading": generate_portfolio(1, df, specialized=True),  # Trading : 1 an
    "Profit": generate_portfolio(0, df, profit=True),
    "Court Terme": generate_portfolio(3, df),
    "Moyen Terme": generate_portfolio(5, df),
    "Long Terme": generate_portfolio(10, df)
}

import re
import pandas as pd

# Fonction pour nettoyer les pourcentages
def clean_percentage(value):
    """
    Nettoie les valeurs de pourcentage en supprimant les caractères non numériques,
    puis les convertit en float.
    """
    try:
        if isinstance(value, str):
            cleaned_value = re.sub(r'[^0-9.\-]', '', value)
            return float(cleaned_value)
        return float(value)
    except ValueError:
        return 0.0

# Nettoyage des colonnes "Dividend Yield" et "Target Potential" (conversion en float)
columns_to_clean = ["Dividend Yield", "Target Potential"]
for col in columns_to_clean:
    if col in df.columns:
        df[col] = df[col].apply(clean_percentage).round(2)

# Ajouter des hyperliens pour chaque ticker
if "Ticker" in df.columns:
    df["Boursorama Link"] = df["Ticker"].apply(lambda ticker: f"https://www.boursorama.com/cours/{ticker.lower()}/")
    df["Yahoo Link"] = df["Ticker"].apply(lambda ticker: f"https://finance.yahoo.com/quote/{ticker}/chart?p={ticker}&range=MAX")

# Renommer les colonnes pour ajouter (%)
if "Target Potential" in df.columns:
    df.rename(columns={"Target Potential": "Target Potential (%)"}, inplace=True)
if "Dividend Yield" in df.columns:
    df.rename(columns={"Dividend Yield": "Dividend Yield (%)"}, inplace=True)

# Supprimer les doublons éventuels après renommage
df = df.loc[:, ~df.columns.duplicated()]

# Colonnes utiles pour les feuilles finales
useful_columns = [
    "Ticker", "Name", "Sector", "Dividend Yield (%)", "Target Potential (%)", "Final Score",
    "Boursorama Link", "Yahoo Link"
]

# Ajouter les colonnes manquantes si nécessaire
def ensure_columns_exist(df, columns):
    for col in columns:
        if col not in df.columns:
            df[col] = ""  # Ajouter une colonne vide si elle est absente

# Exporter les résultats dans un fichier Excel
output_file = f"Portfolio-SP500_{current_date}.xlsx"

with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
    # Exporter les portefeuilles existants
    for horizon, portfolio in portfolios.items():
        # Ajouter les colonnes de liens et renommer si nécessaire
        ensure_columns_exist(portfolio, ["Ticker", "Dividend Yield", "Target Potential"])

        if "Ticker" in portfolio.columns and "Boursorama Link" not in portfolio.columns:
            portfolio["Boursorama Link"] = portfolio["Ticker"].apply(lambda ticker: f"https://www.boursorama.com/cours/{ticker.lower()}/")
        if "Ticker" in portfolio.columns and "Yahoo Link" not in portfolio.columns:
            portfolio["Yahoo Link"] = portfolio["Ticker"].apply(lambda ticker: f"https://finance.yahoo.com/quote/{ticker}/chart?p={ticker}&range=MAX")

        if "Target Potential" in portfolio.columns:
            portfolio.rename(columns={"Target Potential": "Target Potential (%)"}, inplace=True)
        if "Dividend Yield" in portfolio.columns:
            portfolio.rename(columns={"Dividend Yield": "Dividend Yield (%)"}, inplace=True)

        # Arrondir les colonnes pertinentes
        for col in ["Dividend Yield (%)", "Target Potential (%)"]:
            if col in portfolio.columns:
                portfolio[col] = portfolio[col].round(2)

        # Supprimer les doublons éventuels après renommage
        portfolio = portfolio.loc[:, ~portfolio.columns.duplicated()]

        available_columns = [col for col in useful_columns + ["Weight"] if col in portfolio.columns]
        portfolio = portfolio[available_columns]
        portfolio.to_excel(writer, sheet_name=horizon, index=False, startrow=0)

        # Récupérer le workbook et worksheet
        workbook = writer.book
        worksheet = writer.sheets[horizon]
        link_format = workbook.add_format({'color': 'blue', 'underline': 1})

        # Ajouter des liens cliquables
        for row_num, (_, row) in enumerate(portfolio.iterrows(), start=1):
            if "Boursorama Link" in portfolio.columns:
                worksheet.write_url(
                    row_num, portfolio.columns.get_loc("Boursorama Link"),
                    row["Boursorama Link"], link_format, "Boursorama"
                )
            if "Yahoo Link" in portfolio.columns:
                worksheet.write_url(
                    row_num, portfolio.columns.get_loc("Yahoo Link"),
                    row["Yahoo Link"], link_format, "Yahoo Finance"
                )

print(f"Fichier exporté avec succès dans '{output_file}'.")
adjust_column_width(output_file)

################################################################################################################

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from datetime import datetime
import sys

# Configuration SMTP
smtp_server = "smtp.gmail.com"
port = 465
sender_email = "taugourdea@cy-tech.fr"
password = "merz kpoh hcwm xjof"  # Utilisez un mot de passe d'application pour sécuriser l'accès

# Fichiers
current_date = datetime.now().strftime("%Y-%m-%d")
file_1 = f"SP500_Ranking_{current_date}.xlsx"
file_2 = f"Portfolio-SP500_{current_date}.xlsx"
image_files = [
    "Image/logo.png",
    "Image/stock.png",
    "Image/qrcode.png"
]

# Vérification des fichiers
if not os.path.exists(file_1) or not os.path.exists(file_2):
    print(f"🚨 Un ou plusieurs fichiers sont introuvables : {file_1}, {file_2}")
    sys.exit(1)

# Création de l'objet MIMEMultipart pour le message
msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = receiver_email
msg['Subject'] = f"📈 FinanceFinder : Les Meilleures Stratégies d'Investissement pour CTO [{current_date}]"

# Corps de l'email
email_body = f"""Bonjour cher investisseur,

Nous avons le plaisir de vous présenter les meilleures stratégies d'investissement ainsi que les meilleurs actions du marché pour CTO. Vous trouverez le Google Sheet ci-joint :

1. 📊 Classement S&P 500 : Une analyse détaillée des actions, classées selon leur rendement, potentiel de croissance et gain total sur un an
2. 📋 Portefeuille Optimisé : Un portefeuille ajusté à vos priorités d'investissement

Aperçu des Stratégies Proposées
- 📈 Trading Court Terme (1 an) : Sélection d'actions prometteuses basées sur leur potentiel de croissance rapide.
- ⚖️ Profit Équilibré : Diversification entre dividendes réguliers et forte croissance cible.
- 🔄 Court Terme (3 ans) : Portefeuille équilibré entre rendement, volatilité modérée et dividendes attractifs.
- 🏦 Moyen Terme (5 ans) : Actions stables offrant un bon équilibre entre rendement et risque.
- 🌱 Long Terme (10 ans) : Investissements durables misant sur la stabilité et une croissance soutenue.

Ces documents ont été créés pour vous fournir des outils pratiques et des perspectives claires pour vos décisions financières. 

Si vous avez des questions ou souhaitez un suivi, notre équipe se tient à votre disposition.

Cordialement,

💼 L'équipe FinanceFinder
N'hésitez pas à faire un don via PayPal à l'adresse romtaug@gmail.com si cela vous a aidé. 
Accédez à notre outil pour vérifier les signaux d'achats ici : https://romtaugs.shinyapps.io/FinanceFinder/"""

msg.attach(MIMEText(email_body, 'plain'))

# Fonction pour attacher un fichier
def attach_file(msg, file_path):
    try:
        with open(file_path, "rb") as file:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(file.read())
        encoders.encode_base64(attachment)
        filename = os.path.basename(file_path)
        attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(attachment)
        print(f"📎 Fichier joint : {filename}")
    except FileNotFoundError:
        print(f"❌ Erreur : Le fichier {file_path} est introuvable.")

# Attacher les fichiers Excel
attach_file(msg, file_1)
attach_file(msg, file_2)

# Attacher les images
for image in image_files:
    attach_file(msg, image)

# Envoi de l'email
try:
    server = smtplib.SMTP_SSL(smtp_server, port)
    server.login(sender_email, password)
    server.send_message(msg)
    print(f"✅ Email envoyé avec succès à {receiver_email}")
except Exception as e:
    print(f"❌ Erreur lors de l'envoi de l'email : {e}")
finally:
    server.quit()

###################################################################
# On déplace dans les bons dossiers
import os
import shutil
from datetime import datetime

# Récupération de la date actuelle pour formater les noms de fichiers
current_date = datetime.now().strftime("%Y-%m-%d")

# Définir les fichiers et les répertoires cibles avec extension .xlsx
files_and_destinations = {
    f"Data-SP500_{current_date}.xlsx": "Data",
    f"Analysis-SP500_{current_date}.xlsx": "Analysis",
    f"Portfolio-SP500_{current_date}.xlsx": "Portfolio",
    f"SP500_Ranking_{current_date}.xlsx": "Ranking"
}

# Déplacement des fichiers
for file_name, destination_dir in files_and_destinations.items():
    # Vérifier si le fichier existe
    if os.path.exists(file_name):
        # Créer le répertoire cible s'il n'existe pas
        os.makedirs(destination_dir, exist_ok=True)
        # Déplacer le fichier
        shutil.move(file_name, os.path.join(destination_dir, file_name))
        print(f"✅ {file_name} déplacé vers {destination_dir}/")
    else:
        print(f"❌ {file_name} introuvable.")

